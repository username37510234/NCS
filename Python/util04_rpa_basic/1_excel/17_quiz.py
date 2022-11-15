from openpyxl import Workbook
wb = Workbook()
ws = wb.active

ws.append(["학번","출석","퀴즈1","퀴즈2","중간고사","기말고사","프로젝트"])

ws.append([1,10,8,5,14,26,12])
ws.append([2,7,3,7,15,24,18])
ws.append([3,9,5,8,8,12,4])
ws.append([4,7,8,7,17,21,18])
ws.append([5,7,8,7,16,25,15])
ws.append([6,3,5,8,8,17,0])
ws.append([7,4,9,10,16,27,18])
ws.append([8,6,6,6,15,19,17])
ws.append([9,10,10,9,19,30,19])
ws.append([10,9,8,8,20,25,20])

for row in ws.rows:
    for cell in row:
        if cell.column ==4:
            if cell.row == 1:
                continue
            else:
                cell.value = 10
                
ws["H1"] = "총점"
ws["I1"] = "성적"

for i in range(2, ws.max_row + 1):
    k = str(i)
    sum = "=SUM(B"+k+":G"+k+")"
    ws["H"+k] = sum
                

for i in range(2, ws.max_row + 1):
    k = str(i)
    if(ws["B"+k].value<5):
        ws["I"+k] = "F"
    else:
        grade = ("=IF(H"+k+">=90,\"A\",IF(H"+k+">=80,\"B\",IF(H"+k+">=70,\"C\",\"D\")))")
        ws["I"+k] = grade

wb.save("scores.xlsx")