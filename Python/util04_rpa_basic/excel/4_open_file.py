from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")
ws = wb.active

for x in range(1,11):
    for y in range(1,11):
        print(ws.cell(column=y,row=x).value, end=" ")
    print()
    
for x in range(1,ws.max_row):
    for y in range(1,ws.max_column):
        print(ws.cell(row=x,column=y).value, end=" ")
    print()
    