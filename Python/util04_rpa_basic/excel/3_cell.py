from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active
ws.title = "OraenoSheet"

ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"])
print(ws["A1"].value)
print(ws["A10"].value)

print(ws.cell(row=1,column=1).value)
print(ws.cell(row=1,column=2).value)

c = ws.cell(column=3,row=1,value=10)
print(c.value)

for x in range(1,11):
    for g in range(1,11):
        ws.cell(row=x,column=g,value=randint(0,100))

wb.save("sample.xlsx")